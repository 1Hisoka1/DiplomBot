import telebot
from telebot import types
from bs4 import BeautifulSoup
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
import requests
import re
import wget
import os.path
import pyodbc
from openpyxl import load_workbook

bot = telebot.TeleBot('', parse_mode=None)
#база данных
conn_to_db = pyodbc.connect()
cursor = conn_to_db.cursor()
workbook = load_workbook(filename='Лист Microsoft Excel.xlsx', data_only=True)
sheet = workbook['Реконструкиця']



# общая клавиатура
markupotvet = types.ReplyKeyboardMarkup()
btnotv1 = types.KeyboardButton("📝Расписание занятий📝")
btnotv2 = types.KeyboardButton("🗓Учебный план🗓")
btnotv3 = types.KeyboardButton("🎭Мероприятия🎭")
btnotv4 = types.KeyboardButton("📚Библиотека📚")
btnotv5 = types.KeyboardButton("🏢Общежитие🏢")
btnotv6 = types.KeyboardButton("🏥Поликлиника🏥")
btnotv7 = types.KeyboardButton("Рассылка")
markupotvet.add(btnotv5, btnotv2, btnotv3)
markupotvet.add(btnotv1)
markupotvet.add(btnotv6, btnotv7, btnotv4)
# Клава расписания
markupraspis = types.ReplyKeyboardMarkup(resize_keyboard=True)
btnsvsex = types.KeyboardButton('🧾Расписание всех курсов🧾')
btnsodin = types.KeyboardButton('✏️Расписание по группе✏️')
btnzvonok = types.KeyboardButton('🔔Расписание звонков🔔')
btnback = types.KeyboardButton('Назад')
markupraspis.add(btnsvsex)
markupraspis.add(btnback, btnsodin, btnzvonok)

markuprassilka = types.ReplyKeyboardMarkup(resize_keyboard=True)
btnsendphoto = types.KeyboardButton('Отправить сообщение')
btnsendmessage = types.KeyboardButton('Отправить фото')
markuprassilka.add(btnsendmessage, btnsendphoto)
markuprassilka.add(btnback)

markuprgr = types.ReplyKeyboardMarkup(resize_keyboard=True)
btnsvsex = types.KeyboardButton('Выберите группу')
markuprgr.add(btnsvsex)
url = "http://kktd.ru/studentu/raspisanie/"

response = requests.get(url)
bs = BeautifulSoup(response.text, "html.parser")
temp = bs.find_all(class_='doc_link')[0]
temp1 = bs.find_all(class_='doc_link')[1]
sravnperv = bs.find_all(class_='doc_link')[0]
sravnstarsh= bs.find_all(class_='doc_link')[1]
starsh = sravnstarsh.text
perv = sravnperv.text
linkperv = str(temp.get('href'))
linkstarsh = str(temp1.get('href'))
linkredp = linkperv.rstrip('/').split('/')[-1]
linkreds = linkstarsh.rstrip('/').split('/')[-1]
if os.path.isfile(linkreds):
    print('Уже есть это расписание')
else:
    wget.download(linkstarsh)
if os.path.isfile(linkredp):
    print('уже есть это расписание')
else:
    wget.download(linkperv)

workbook = load_workbook(filename='Лист Microsoft Excel.xlsx', data_only=True)
sheet = workbook['Реконструкиця']



def rasp09B():
    f = open('По группам.txt', 'w')
    a = 0
    for item in sheet:
        a += 1
    for i in range(1, a+1):
        b = sheet['A'+str(i)].value
        c = sheet['B' + str(i)].value
        d = sheet['C' + str(i)].value
        if b != None:
            f.write('\n')
            f.write(b + '\n')
        if c != None and d!=None:
            f.write(c)
            f.write(' ')
            f.write(str(d))
            f.write('\n')
        if c != None and d == None:
                f.write(c + '\n')
def rasp09K():
    a = 0
    f = open('По группам.txt', 'w')
    for item in sheet:
        a += 1
    for i in range(1, a+1):
        b = sheet['A'+str(i)].value
        c = sheet['D' + str(i)].value
        d = sheet['E' + str(i)].value
        if b != None:
            f.write('\n')
            f.write(b + '\n')
        if c != None and d!=None:
            f.write(c)
            f.write(' ')
            f.write(str(d))
            f.write('\n')
        if c != None and d == None:
                f.write(c + '\n')


@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    if message.text == '/start':
        bot.send_message(message.chat.id, 'Привет, Введи своё имя и после выбери группу')
#извлекаем следующее сообщение для имени
        bot.register_next_step_handler(message, add_user)
    if message.text == '/help':
        bot.send_message(message.chat.id, 'Здесь ты можешь узнать своё расписание\nКуда обратится если ты вдруг заболел\nУзнать актуальное расписание звонков\nУзнать мероприятия, которые проходят\nВопросы по общежитию')
#Добавляем юсера
def add_user(message):
    user_id = message.chat.id
    try:
        cursor.execute('SELECT * FROM users WHERE user_id=?', (user_id))
        cursor.execute("INSERT INTO users (user_id, user_name) VALUES ({}, '{}')".format(user_id, str(message.text)))
        cursor.commit()
        bot.send_message(message.chat.id, 'Вы успешно зарегестрировались!', reply_markup=markuprgr)
        bot.register_next_step_handler(message, groups)
    except:
        bot.send_message(message.chat.id, 'У вас уже был создан аккаунт, продолжим', reply_markup=markupotvet)
def group_mar():
    markupgrupp = InlineKeyboardMarkup()
    markupgrupp.row_width = 2
    # Длина столбцов для ццикла клави
    group = cursor.execute('SELECT COUNT(group_id) FROM groups').fetchone()[0]
    # Значения для цикла
    groupcik = cursor.execute('SELECT group_id FROM groups').fetchall()
    for i in range(group):
        call1 = 'callgroup'
        call1 = call1+str(i)
        markupgrupp.add(InlineKeyboardButton(*groupcik[i], callback_data=call1))
    return markupgrupp

def groups(message):
    bot.send_message(message.chat.id, 'Выберите группу', reply_markup=group_mar())


@bot.message_handler(content_types=['text', 'document', 'photo'])
def markup_otv(message):
    # удобство
    ChatID = message.chat.id
    a = message.chat.id
    group = cursor.execute(f'SELECT group_id FROM group_member WHERE user_id = {a}').fetchall()
    gr = ''.join(group[0])
    if message.text =='Рассылка':
        if ChatID == 1089155407:
            bot.send_message(message.chat.id, 'Выберите что хотите отправить:', reply_markup=markuprassilka)
        else:
            bot.send_message(message.chat.id, 'У вас недостаточно прав')

    if ChatID == 1089155407 and message.text == 'Отправить сообщение':
        bot.send_message(message.chat.id, 'Напишите сообщение для отправки')
        bot.register_next_step_handler(message, sendrastext)

    if ChatID == 1089155407 and message.text == 'Отправить фото':
        bot.send_message(message.chat.id, 'Выберите фото для отправки')
        bot.register_next_step_handler(message, sendrasphoto)

    if (ChatID != 1089155407 and message.text == 'Отправить сообщение') or (ChatID != 1089155407 and message.text == 'Отправить фото'):
        bot.send_message(message.chat.id, 'У вас недостаточно прав')

    if message.text == '📝Расписание занятий📝':
        bot.send_message(ChatID, 'Выберте кнопку которая вам нужна', reply_markup=markupraspis)
    if gr == '09.02.07-4':
        rasp09B()
    if gr == '09.02.07-4К':
        rasp09K()
    if message.text == '🧾Расписание всех курсов🧾':
        bot.send_message(ChatID, 'Первокурсники')
        rasperv = open(r'C:\\Users\\goodg\\PycharmProjects\\diplom\\'+linkredp, 'rb')
        bot.send_document(ChatID, rasperv)
        bot.send_message(ChatID, 'Старшекурсники')
        raspstarsh = open(r'C:\\Users\\goodg\\PycharmProjects\\diplom\\'+linkreds, 'rb')
        bot.send_document(ChatID, raspstarsh)
    if message.text == '🔔Расписание звонков🔔':
        bot.send_message(ChatID, 'Выберите на какой день недели', reply_markup=gen_markup())
    if message.text == 'Назад':
        bot.send_message(message.chat.id, 'Вы вернулись в главное меню', reply_markup=markupotvet)
    if message.text == '🏢Общежитие🏢':
        with open("Общежитие.txt", "rb") as f:  # открываем документ
            contents = f.read().decode("UTF-8")  # считываем все строки
            bot.send_message(message.chat.id, contents)  # отправляем содержимое документа
    if message.text == '🏥Поликлиника🏥':
        with open("Поликлинника.txt", "rb") as f:  # открываем документ
            contents = f.read().decode("UTF-8")  # считываем все строки
            bot.send_message(message.chat.id, contents)  # отправляем содержимое документа
    if message.text == '🗓Учебный план🗓':
        bot.send_message(message.chat.id, '🗒Учебный план всех курсов🗒 https://disk.yandex.ru/d/YPynT_gSKPipzQ')
    if message.text == '📚Библиотека📚':
        with open("Библиотека.txt", "rb") as f:  # открываем документ
            contents = f.read().decode("UTF-8")  # считываем все строки
            bot.send_message(message.chat.id, contents)  # отправляем содержимое документа
    if message.text == '✏️Расписание по группе✏️':
        a = message.chat.id
        group = cursor.execute(f'SELECT group_id FROM group_member WHERE user_id = {a}').fetchall()
        gr = ''.join(group[0])
        if gr == '09.02.07-4':
            with open("по группам.txt", "rb") as f:  # открываем документ
                contents = f.read().decode("Windows-1251")  # считываем все строки
                bot.send_message(message.chat.id, contents)  # отправляем содержимое документа
        if gr == '09.02.07-4К':
            with open("по группам.txt", "rb") as f:  # открываем документ
                contents = f.read().decode("Windows-1251")  # считываем все строки
                bot.send_message(message.chat.id, contents)  # отправляем содержимое документа
    if message.text == '🎭Мероприятия🎭':
        with open("Меро.txt", "rb") as f:  # открываем чдокумент
            contents = f.read().decode("UTF-8")  # считываем все строки
            bot.send_message(message.chat.id, contents)  # отправляем содержимое документа

def sendrastext(message):
    rass = cursor.execute('SELECT user_id FROM users').fetchall()
    for i in range(len(rass)):
        ra = ''.join(str(rass[i]))
        raa = re.sub("[]|[|,|(|)| ]", "", str(ra))
        bot.send_message(raa, message.text)
    try:
        bot.send_message(message.chat.id, 'Если сообщение отправилось вам тоже, то рассылка прошла успешно')
    except telebot.apihelper.ApiException:
        pass
def sendrasphoto(message):
    photo = max(message.photo, key=lambda x: x.height)
    rass = cursor.execute('SELECT user_id FROM users').fetchall()
    for i in range(len(rass)):
        ra = ''.join(str(rass[i]))
        raa = re.sub("[]|[|,|(|)| ]", "", str(ra))
        photo = max(message.photo, key=lambda x: x.height)
        bot.send_photo(raa, photo.file_id)
    try:
        bot.send_message(message.chat.id, 'Если фото отправилось вам тоже, то рассылка прошла успешно')
    except telebot.apihelper.ApiException:
        pass

def gen_markup():
    markup = InlineKeyboardMarkup()
    markup.row_width = 2
    markup.add(InlineKeyboardButton("Понедельник", callback_data="pn"))
    markup.add(InlineKeyboardButton("Вторник-пятница", callback_data="vt_pt"))
    markup.add(InlineKeyboardButton('Суббота', callback_data="cybb"))
    return markup
#Инлайн кнопки для групп
@bot.callback_query_handler(func=lambda call: True)
def callback_query(call):
    photpn = open('звонки_пн.jpg', 'rb')
    photvtpt = open(r'звонки_вт-пт.jpg', 'rb')
    photcybb = open(r'звонки_сб.jpg', 'rb')

    call_gr = call.message.chat.id
    bot.answer_callback_query(callback_query_id=call.id)
    if call.data == "pn":
        bot.send_photo(call.message.chat.id, photo=photpn)
    if call.data == "vt_pt":
        bot.send_photo(call.message.chat.id, photo=photvtpt)
    if call.data == "cybb":
        bot.send_photo(call.message.chat.id, photo=photcybb)
    group = cursor.execute('SELECT COUNT(group_id) FROM groups').fetchone()[0]
    # Значения для цикла
    groupcik = cursor.execute('SELECT group_id FROM groups').fetchall()
    for i in range(group):
        rasp09B()
        call1 = 'callgroup'
        call1 = call1+str(i)
        if call.data == call1:
            gr = ''.join(groupcik[i])
            print(gr)
            cursor.execute('INSERT INTO group_member(user_id, group_id) VALUES(?, ?)', (call_gr, gr))
            cursor.commit()
            bot.send_message(call.message.chat.id, 'Отлично, продолжим', reply_markup=markupotvet)



bot.infinity_polling(none_stop=True)